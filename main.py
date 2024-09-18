import streamlit as st
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier
from sklearn.pipeline import Pipeline
from sklearn.utils.multiclass import type_of_target
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

def load_data(file):
    df = pd.read_excel(file)
    return df

def train_model(df):
    st.write("Nama kolom yang ada:", df.columns)
    
    required_columns = [
        'Test Condition\n(Normal/Negative/Abnormal}', 'Scenario', 'Remark', 'Version',
        'Test Script ID', 'Test Case Description', 'Steps Name', 'Steps Description', 'Expected Result'
    ]
    
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        st.error(f"Kolom berikut tidak ditemukan dalam dataset: {missing_columns}")
        return None
    
    X = df[['Menu', 'Function Name']].astype(str)
    y = df[required_columns]

    X_combined = X.apply(lambda x: ' '.join(x), axis=1)
 
    combined = pd.concat([X_combined, y], axis=1)
    combined = combined.dropna()

    X_combined = combined.iloc[:, 0]
    y = combined.iloc[:, 1:]
    
    for column in y.columns:
        y[column] = y[column].astype(str).astype('category')
        if type_of_target(y[column]) not in ['binary', 'multiclass']:
            raise ValueError(f"Target column {column} is not suitable for classification: {type_of_target(y[column])}")
    
    model = Pipeline([
        ('tfidf', TfidfVectorizer()),
        ('clf', RandomForestClassifier())
    ])
    
    models = {}
    for column in y.columns:
        model.fit(X_combined, y[column])
        models[column] = model
    
    return models

def predict_template_fields(models, menu, function_name):
    input_data = ' '.join([menu, function_name])
    predictions = {}
    for column, model in models.items():
        prediction = model.predict([input_data])[0]
        predictions[column] = prediction
    return predictions

def get_data_for_input(df, menus, function_names):
    df['Menu'] = df['Menu'].str.strip().str.lower()
    df['Function Name'] = df['Function Name'].str.strip().str.lower()
    
    filtered_data = df[
        df['Menu'].isin([menu.strip().lower() for menu in menus]) & 
        df['Function Name'].isin([func.strip().lower() for func in function_names])
    ]
    
    if filtered_data.empty:
        st.error("No data found for the given inputs.")
        st.write("Filtered data (empty):", filtered_data)
        return None
    else:
        st.write("Filtered data found:", filtered_data)
    
    return filtered_data

def generate_excel(filtered_data, predictions):
    new_df = filtered_data.copy()
    
    for column in predictions:
        new_df[column] = predictions[column]

    output_file = f"{filtered_data.iloc[0]['Menu']}_{filtered_data.iloc[0]['Function Name']}_test_script.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Sheet1'
    
    for row in dataframe_to_rows(new_df, index=False, header=True):
        worksheet.append(row)
    
    for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(size=12)
    
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width
    
    workbook.save(output_file)
    
    return output_file

def main():
    st.title("TS Generator Audit Excel Workbook")
    
    uploaded_file = st.file_uploader("Upload your dataset (Excel file)", type=["xlsx"])
    
    if uploaded_file is not None:
        template_df = load_data(uploaded_file)
        st.success("Dataset loaded successfully!")

        models = train_model(template_df)
        if models:
            st.success("Model trained successfully!")

            menu_inputs = []
            function_name_inputs = []

            if 'menu_inputs' not in st.session_state:
                st.session_state.menu_inputs = []
            if 'function_name_inputs' not in st.session_state:
                st.session_state.function_name_inputs = []

            with st.form(key='input_form'):
                menu = st.text_input("Enter Menu")
                function_name = st.text_input("Enter Function Name")
                
                if st.form_submit_button("Add"):
                    if menu and function_name:
                        st.session_state.menu_inputs.append(menu)
                        st.session_state.function_name_inputs.append(function_name)
                    else:
                        st.error("Please enter both Menu and Function Name")

            st.write("Current inputs:")
            for i, (menu, function_name) in enumerate(zip(st.session_state.menu_inputs, st.session_state.function_name_inputs)):
                st.write(f"{i+1}. Menu: {menu}, Function Name: {function_name}")
            
            if st.button("Generate Excel"):
                if st.session_state.menu_inputs and st.session_state.function_name_inputs:
                    filtered_data = get_data_for_input(template_df, st.session_state.menu_inputs, st.session_state.function_name_inputs)
                    if filtered_data is not None:
                        predictions = {}
                        for menu, function_name in zip(st.session_state.menu_inputs, st.session_state.function_name_inputs):
                            prediction = predict_template_fields(models, menu, function_name)
                            predictions.update(prediction)
                        # Ensure that only unique entries are kept for Excel generation
                        if not filtered_data.empty:
                            output_file = generate_excel(filtered_data, predictions)
                            st.success(f"Excel file generated: {output_file}")
                            with open(output_file, 'rb') as file:
                                st.download_button(
                                    label="Download generated Excel file",
                                    data=file,
                                    file_name=output_file,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                else:
                    st.error("Please add at least one Menu and Function Name")

if __name__ == "__main__":
    main()
